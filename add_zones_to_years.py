import argparse
import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Dict

try:
    from openpyxl import load_workbook
except ImportError:
    pass


def normalize_text(value: str) -> str:
    """Normalize text for robust matching."""
    value = value.strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("'", " ")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def extract_commune_name(prefixed_label: str) -> str:
    """Extract the plain commune name from a prefixed label like '1003 - BIR EL DJIR'."""
    match = re.match(r"^\d+\s*-\s*(.+)$", prefixed_label.strip())
    return match.group(1).strip() if match else prefixed_label.strip()


def load_zone_mapping(zones_csv: Path) -> Dict[str, str]:
    """Load zones_clean_prefixed.csv and map normalized commune names to zones."""
    mapping = {}
    with zones_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            commune_name = row.get("commune_name", "").strip()
            zone = row.get("zone", "").strip()
            if commune_name and zone:
                # Normalize the commune name (extract base name if it has a prefix)
                base_name = extract_commune_name(commune_name)
                key = normalize_text(base_name)
                # Keep the most recent/frequent entry if there are duplicates
                if key not in mapping:
                    mapping[key] = zone
    return mapping


def resolve_zone(commune_name: str, zone_mapping: Dict[str, str]) -> str:
    """Resolve a commune name to its zone using fuzzy matching."""
    key = normalize_text(commune_name)
    if key in zone_mapping:
        return zone_mapping[key]
    return ""


def process_workbook(workbook_path: Path, year: str, zone_mapping: Dict[str, str], output_csv: Path):
    """Extract data from workbook, add zone column, and write to CSV."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"Error: openpyxl not available for {year}")
        return

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    # Find relevant column indices
    try:
        commune_idx = headers.index("COMMUNE_NAME")
    except ValueError:
        print(f"Error: COMMUNE_NAME column not found in {workbook_path}")
        return

    output_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        commune_name = row[commune_idx]
        if not isinstance(commune_name, str) or not commune_name.strip():
            continue

        # Extract base name and resolve zone
        base_name = extract_commune_name(commune_name)
        zone = resolve_zone(base_name, zone_mapping)

        # Build output row with original data plus zone
        output_row = dict(zip(headers, row))
        output_row["zone"] = zone
        output_rows.append(output_row)

    # Write output
    if output_rows:
        output_headers = headers + ["zone"]
        with output_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=output_headers)
            writer.writeheader()
            for out_row in output_rows:
                writer.writerow(out_row)
        print(f"Wrote {len(output_rows)} rows to {output_csv}")
    else:
        print(f"No rows written to {output_csv}")


def main():
    parser = argparse.ArgumentParser(description="Add zone column to insurance workbooks for 2023, 2024, 2025.")
    parser.add_argument(
        "--zones",
        default="zones_clean_prefixed.csv",
        help="Path to zones_clean_prefixed.csv for zone mapping.",
    )
    parser.add_argument(
        "--workbook-2023",
        default="cleaned_insurance_sheet 2023.xlsx",
        help="Path to 2023 workbook.",
    )
    parser.add_argument(
        "--workbook-2024",
        default="cleaned_catnat_2024 - Sheet1.xlsx",
        help="Path to 2024 workbook.",
    )
    parser.add_argument(
        "--workbook-2025",
        default="cleaned_catnat_2025 - Sheet1.xlsx",
        help="Path to 2025 workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Output directory for generated CSVs.",
    )

    args = parser.parse_args()
    zones_path = Path(args.zones)
    output_dir = Path(args.output_dir)

    # Load zone mapping once
    print("Loading zone mapping...")
    zone_mapping = load_zone_mapping(zones_path)
    print(f"Loaded {len(zone_mapping)} zone mappings")

    # Process each year
    for year, workbook_arg in [
        ("2023", "workbook_2023"),
        ("2024", "workbook_2024"),
        ("2025", "workbook_2025"),
    ]:
        workbook_path = Path(getattr(args, workbook_arg))
        output_csv = output_dir / f"cleaned_{year}_with_zones.csv"
        if workbook_path.exists():
            print(f"\nProcessing {year}...")
            process_workbook(workbook_path, year, zone_mapping, output_csv)
        else:
            print(f"Warning: {workbook_path} not found")


if __name__ == "__main__":
    main()
