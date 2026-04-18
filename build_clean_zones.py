import argparse
import csv
import difflib
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set, Tuple


def normalize_text(value: str) -> str:
    """Normalize text for robust commune matching across data sources."""
    value = value.strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("'", " ")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def compact_text(value: str) -> str:
    return normalize_text(value).replace(" ", "")


def to_ascii(value: str) -> str:
    """Convert accented text to plain ASCII while preserving readable names."""
    folded = unicodedata.normalize("NFKD", value)
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    folded = folded.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", folded).strip()


def to_upper_ascii(value: str) -> str:
    return to_ascii(value).upper()


def parse_group_letter(spec: str) -> str:
    match = re.search(r"groupe\s*([abc])", spec, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def parse_referenced_group_letters(spec: str) -> Set[str]:
    # Example: "... groupe de communes B et C" -> {"B", "C"}
    letters = set(re.findall(r"\b([ABC])\b", spec.upper()))
    letters.discard("A")
    return letters


def parse_commune_list(spec: str) -> List[str]:
    if ":" in spec:
        payload = spec.split(":", 1)[1]
    else:
        payload = spec
    payload = payload.strip().strip(".")
    if not payload:
        return []
    parts = [part.strip().strip(".") for part in payload.split(",")]
    return [part for part in parts if part]


def best_match_commune(
    token: str,
    normalized_to_canonical: Dict[str, str],
    compact_to_canonical: Dict[str, str],
) -> str:
    norm = normalize_text(token)
    compact = compact_text(token)
    if not norm:
        return ""
    if norm in normalized_to_canonical:
        return normalized_to_canonical[norm]
    if compact in compact_to_canonical:
        return compact_to_canonical[compact]

    candidates = difflib.get_close_matches(norm, normalized_to_canonical.keys(), n=1, cutoff=0.75)
    if candidates:
        return normalized_to_canonical[candidates[0]]

    # Fallback for minor spacing differences and truncated names.
    contains = [k for k in normalized_to_canonical if norm in k or k in norm]
    if len(contains) == 1:
        return normalized_to_canonical[contains[0]]

    compact_contains = [k for k in compact_to_canonical if compact in k or k in compact]
    if len(compact_contains) == 1:
        return compact_to_canonical[compact_contains[0]]

    return ""


def match_communes_from_token(
    token: str,
    normalized_to_canonical: Dict[str, str],
    compact_to_canonical: Dict[str, str],
) -> Set[str]:
    single = best_match_commune(token, normalized_to_canonical, compact_to_canonical)
    if single:
        return {single}

    token_compact = compact_text(token)
    if not token_compact:
        return set()

    matches = set()
    for key, commune in compact_to_canonical.items():
        if len(key) < 5:
            continue
        if key in token_compact:
            matches.add(commune)
            continue
        if key.startswith("el") and key[2:] and key[2:] in token_compact:
            matches.add(commune)
            continue
        if key.startswith("al") and key[2:] and key[2:] in token_compact:
            matches.add(commune)

    return matches


def load_cities(cities_csv: Path):
    wilaya_info: Dict[str, str] = {}
    wilaya_communes: Dict[str, Set[str]] = defaultdict(set)

    with cities_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row["wilaya_code"].strip().zfill(2)
            wilaya_info[code] = row["wilaya_name"].strip()
            wilaya_communes[code].add(row["commune_name"].strip())

    normalized_lookup: Dict[str, Dict[str, str]] = {}
    compact_lookup: Dict[str, Dict[str, str]] = {}
    for code, communes in wilaya_communes.items():
        lookup = {}
        compact_map = {}
        for commune in communes:
            lookup[normalize_text(commune)] = commune
            compact_map[compact_text(commune)] = commune
        normalized_lookup[code] = lookup
        compact_lookup[code] = compact_map

    return wilaya_info, wilaya_communes, normalized_lookup, compact_lookup


def load_commune_labels(workbook_paths: List[Path]):
    """Load canonical commune labels from one or more Excel workbooks.

    The workbook stores commune names like '1003 - BIR EL DJIR'. This function
    maps the raw commune name to the exact prefixed uppercase label. When multiple
    workbooks are provided, earlier workbooks have higher priority.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    canonical: Dict[str, str] = {}

    def is_prefixed(label: str) -> bool:
        return bool(re.match(r"^\d+\s*-\s*.+$", label))

    for workbook_path in workbook_paths:
        if not workbook_path or not workbook_path.exists():
            continue

        label_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        commune_idx = headers.index("COMMUNE_NAME")

        for row in ws.iter_rows(min_row=2, values_only=True):
            cell = row[commune_idx]
            if not isinstance(cell, str):
                continue
            label = re.sub(r"\s+", " ", cell.strip().upper())
            if label in {"", "118 -"}:
                continue
            match = re.match(r"^(\d+)\s*-\s*(.+)$", label)
            raw = match.group(2).strip() if match else label
            key = normalize_text(raw)
            if key and is_prefixed(label):
                label_counts[key][label] += 1

        for key, labels in label_counts.items():
            if key not in canonical:
                canonical[key] = sorted(labels.items(), key=lambda item: (-item[1], item[0]))[0][0]

    return canonical


def resolve_commune_label(commune: str, commune_labels: Dict[str, str]) -> str:
    """Resolve a commune name to a prefixed uppercase label.

    Tries exact normalization first, then compact matching, then fuzzy matching
    against available workbook labels. Falls back to uppercase ASCII if no good
    label is found.
    """
    key = normalize_text(commune)
    if key in commune_labels:
        return commune_labels[key]

    compact = compact_text(commune)
    if compact:
        compact_matches = [k for k in commune_labels if compact == compact_text(k)]
        if len(compact_matches) == 1:
            return commune_labels[compact_matches[0]]

    candidates = difflib.get_close_matches(key, list(commune_labels.keys()), n=3, cutoff=0.8)
    if candidates:
        best = candidates[0]
        return commune_labels[best]

    return to_upper_ascii(commune)


def load_zone_rows(zone_csv: Path):
    rows = []
    with zone_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for raw in reader:
            if len(raw) < 4:
                continue
            wilaya_code = raw[0].strip().zfill(2)
            wilaya_name = raw[1].strip()
            spec = raw[2].strip()
            zone = raw[3].strip()
            rows.append(
                {
                    "wilaya_code": wilaya_code,
                    "wilaya_name": wilaya_name,
                    "spec": spec,
                    "zone": zone,
                }
            )
    return rows


def build_clean_dataset(cities_csv: Path, zone_csv: Path, output_csv: Path, workbook_paths: List[Path] | None = None):
    wilaya_info, wilaya_communes, normalized_lookup, compact_lookup = load_cities(cities_csv)
    commune_labels = load_commune_labels(workbook_paths or [])
    zone_rows = load_zone_rows(zone_csv)

    explicit_group_communes: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    unmatched_tokens: List[Tuple[str, str, str]] = []

    for row in zone_rows:
        code = row["wilaya_code"]
        spec = row["spec"]
        group_letter = parse_group_letter(spec)
        if not group_letter:
            continue
        if "toutes les communes" in normalize_text(spec):
            continue

        tokens = parse_commune_list(spec)
        for token in tokens:
            matched_set = match_communes_from_token(
                token,
                normalized_lookup.get(code, {}),
                compact_lookup.get(code, {}),
            )
            if matched_set:
                explicit_group_communes[(code, group_letter)].update(matched_set)
            else:
                unmatched_tokens.append((code, group_letter, token))

    output_rows = set()

    for row in zone_rows:
        code = row["wilaya_code"]
        spec = row["spec"]
        zone = row["zone"]

        if code not in wilaya_communes:
            continue

        spec_norm = normalize_text(spec)
        if spec_norm in {"(non precise)", "non precise"}:
            continue

        all_communes = set(wilaya_communes[code])
        selected_communes: Set[str] = set()

        if "toutes les communes" in spec_norm:
            selected_communes = all_communes
        else:
            group_letter = parse_group_letter(spec)
            if group_letter == "A" and "autres que celles figurant" in spec_norm:
                ref_letters = parse_referenced_group_letters(spec)
                excluded = set()
                for letter in ref_letters:
                    excluded |= explicit_group_communes.get((code, letter), set())
                selected_communes = all_communes - excluded
            else:
                tokens = parse_commune_list(spec)
                for token in tokens:
                    matched_set = match_communes_from_token(
                        token,
                        normalized_lookup.get(code, {}),
                        compact_lookup.get(code, {}),
                    )
                    selected_communes.update(matched_set)

        output_wilaya_name = to_upper_ascii(wilaya_info.get(code, row["wilaya_name"].title()))
        for commune in selected_communes:
            output_commune_name = resolve_commune_label(commune, commune_labels)
            output_rows.add((code, output_wilaya_name, output_commune_name, zone))

    sorted_rows = sorted(output_rows, key=lambda x: (int(x[0]), x[1], x[2], x[3]))

    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["wilaya_code", "wilaya_name", "commune_name", "zone"])
        writer.writerows(sorted_rows)

    if unmatched_tokens:
        print("Unmatched tokens (review recommended):")
        for code, group, token in unmatched_tokens:
            print(f"  wilaya={code}, group={group}, token={token}")

    print(f"Wrote {len(sorted_rows)} rows to {output_csv}")


def main():
    parser = argparse.ArgumentParser(description="Build clean zone/commune mapping from two CSV files.")
    parser.add_argument(
        "--cities",
        default="algeria_cities.csv",
        help="Path to algeria_cities CSV.",
    )
    parser.add_argument(
        "--zones",
        default="Copy of zone aec - Sheet5.csv",
        help="Path to zone CSV.",
    )
    parser.add_argument(
        "--output",
        default="zones_clean.csv",
        help="Output CSV path.",
    )
    parser.add_argument(
        "--workbook",
        default="cleaned_catnat_2024 - Sheet1.xlsx",
        help="Workbook containing canonical prefixed uppercase commune names.",
    )

    args = parser.parse_args()
    workbook_candidates = [
        Path(args.workbook),
        Path("cleaned_catnat_2025 - Sheet1.xlsx"),
        Path("cleaned_insurance_sheet 2023.xlsx"),
    ]
    build_clean_dataset(Path(args.cities), Path(args.zones), Path(args.output), workbook_candidates)


if __name__ == "__main__":
    main()
